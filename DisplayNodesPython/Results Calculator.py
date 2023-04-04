import os
import openpyxl

# Set the directory path for the input files
input_dir = r"C:\Users\Angel\source\repos\VRPdiss\DisplayNodes\Customer nodes\Results\Initial Days"

# Set the directory path for the output file
output_file = r"C:\Users\Angel\source\repos\VRPdiss\DisplayNodes\Customer nodes\Results\Total results\CombinedAgents.xlsx"

# Load the output file or create it if it does not exist
if os.path.exists(output_file):
    wb = openpyxl.load_workbook(output_file)
else:
    wb = openpyxl.Workbook()

# Select the first sheet
ws = wb.active

# Loop through each file in the input directory
for i in range(1, 31):
    filename = f"CombinedAgents{i}.txt"
    file_path = os.path.join(input_dir, filename)

    # Read the data from the file
    with open(file_path, "r") as file:
        data = file.readlines()

    # Initialize a dictionary to store the data for each agent on this day
    agent_data = {}

    # Loop through each line of data and update the corresponding row in the Excel sheet
    for line in data:
        # Check if the line contains "gained" or "saved"
        if "gained" not in line and "saved" not in line:
            continue

        # Split the line into the agent number and value string
        agent_str, value_str = line.split(" has ")

        # Extract the agent number as an integer
        agent_num = int(agent_str.split(" ")[-1])

        # Extract the numerical value from the value string
        value_str = value_str.strip().replace("km", "")
        if "gained" in value_str:
            value_float = -float(value_str.split()[-1])
        else:
            value_float = float(value_str.split()[-1])

        # Add the value to the agent's data for this day
        if agent_num in agent_data:
            agent_data[agent_num] += value_float
        else:
            agent_data[agent_num] = value_float

    # Loop through each agent and update the corresponding row and column in the Excel sheet
    for agent_num in range(1, 31):
        # Update the corresponding row and column in the Excel sheet
        row_num = agent_num + 1
        col_num = i + 1  # Shift the column index to the right
        cell = ws.cell(row=row_num, column=col_num)

        # If the agent has data for this day, update the cell value
        if agent_num in agent_data:
            cell.value = agent_data[agent_num]
        else:
            cell.value = 0

# Add column headers
for i in range(1, 31):
    header_cell = ws.cell(row=1, column=i + 1)  # Shift the column index to the right
    header_cell.value = f"Day {i}"
    
# Insert Agent labels
for i in range(1, 31):
    agent_cell = ws.cell(row=i + 1, column=1)
    agent_cell.value = f"Agent {i}"

# Save the Excel sheet
wb.save(output_file)
